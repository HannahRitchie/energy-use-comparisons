#!/usr/bin/env python3
from openpyxl import load_workbook
import json

# Read the Excel file
wb = load_workbook('Energy use of products.xlsx', data_only=True)
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
product_idx = headers.index('Product')
wattage_idx = headers.index('Energy rating (W)')
user_sel_idx = headers.index('User selection')
hours_idx = headers.index('Hours per day')
wh_idx = headers.index('Energy used (Wh)')

products = []

# Skip header row
for row in ws.iter_rows(min_row=2, values_only=True):
    product_name = row[product_idx]
    wattage = row[wattage_idx]
    user_selection = row[user_sel_idx]
    hours_per_day = row[hours_idx]
    energy_wh = row[wh_idx]

    product = {
        "name": product_name,
        "wattage": None if wattage is None else float(wattage),
        "fixed_wh": None if energy_wh is None else float(energy_wh),
        "input_type": None,
        "default_value": None,
        "label": None
    }

    # Determine input type based on 'User selection' column
    if user_selection is None or user_selection == '':
        product['input_type'] = 'fixed'
        product['default_value'] = None
    elif user_selection == 'Hours':
        product['input_type'] = 'hours'
        product['default_value'] = float(hours_per_day) if hours_per_day is not None else 1.0
        # fixed_wh should already be calculated from the Excel formula
    elif user_selection == 'Number of cycles':
        product['input_type'] = 'cycles'
        product['default_value'] = 1
    elif user_selection == 'Number of full boils':
        product['input_type'] = 'boils'
        product['default_value'] = 1
    elif user_selection == 'Number of queries':
        product['input_type'] = 'queries'
        product['default_value'] = 1
    elif user_selection == 'Miles driven':
        product['input_type'] = 'miles'
        product['default_value'] = 10
    elif user_selection == 'Per day' or user_selection == 'Per day (full charge)':
        product['input_type'] = 'fixed'
        product['default_value'] = None
        product['label'] = 'per day'

    products.append(product)

# Save to JSON file
with open('products_data.json', 'w') as f:
    json.dump(products, f, indent=2)

print(f"Converted {len(products)} products to JSON")
print("\nElectric oven:", [p for p in products if p['name'] == 'Electric oven'][0])
print("Gas oven:", [p for p in products if p['name'] == 'Gas oven'][0])
print("Fridge-freezer:", [p for p in products if p['name'] == 'Fridge-freezer'][0])
print("Charging phone:", [p for p in products if p['name'] == 'Charging a mobile phone'][0])
